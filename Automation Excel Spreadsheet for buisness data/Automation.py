import openpyxl


path ="Client Configuration - Scheduled-03-22-2023(2).xlsx"

# Load the source workbook
wb_obj = openpyxl.load_workbook(path)

# Worksheet is created
wb_obj = openpyxl.load_workbook(path, data_only=True)

sheet_obj = wb_obj.active
m_row = sheet_obj.max_row


# Load the destination workbook
dest_wb = openpyxl.load_workbook("test1.xlsx")

# Loop will print al values
# of first column
for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    print(cell_obj.value) 

for i in range(1, m_row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    dest_wb.active.append([cell_obj.value])
    

dest_wb.save("test1.xlsx")

